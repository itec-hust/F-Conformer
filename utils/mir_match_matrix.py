from datetime import datetime
import mir_eval
import argparse
import openpyxl
from openpyxl.styles import PatternFill
import os
import numpy as np
import librosa
import copy
import pretty_midi
pretty_midi.pretty_midi.MAX_TICK = 1e10
'''
reference https://blog.csdn.net/koudailidexiaolong/article/details/125646925
          https://www.jb51.net/article/252141.htm
'''
# python mir_match_matrix.py --label_dir /home/data/lyq/dataset/OMAP2/aligned_label/aligned_audio_align_221225 --result_dir result_dir
# python /home/data/wrm_data/temp_task/piano_transcription-master/utils/mir_match_matrix.py --label_dir /home/data/wrm_data/temp_task/piano_transcription-master/workspaces/hdf5s/maestro/v3_test_labels --result_dir /home/data/wrm_data/temp_task/piano_transcription-master/workspaces/probs/Four_blocks/none/dataset=maestro/test/batch_size=2/505000_iterations/505000_iterations_midi
SUSTAIN_TOLERANCE = 10  # notes with same pitch in SUSTAIN_TOLERANCE * onset_tolerance
HARMONIC_TOLERANCE = 0  # tolorance of harmonic in pitch range(e.g. 12±1/88)


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--label_dir', '-l', type=str, default="/home/data/wrm_data/temp_task/piano_transcription-master/workspaces/hdf5s/maestro_labels")
    parser.add_argument('--result_dir', '-r', type=str, required=True)
    parser.add_argument('--save_dir', '-s', type=str, default=None)
    parser.add_argument('--onset-win', type=float, default=0.05)
    parser.add_argument('--offset-win', type=float, default=0.05)
    parser.add_argument('--offset', action='store_true', default=False)
    return parser.parse_args()


def extract_labels_from_midi(midi_file):
    midi_data = pretty_midi.PrettyMIDI(midi_file)
    outputs = []
    for instrument in midi_data.instruments:
        notes = instrument.notes
        for note in notes:
            start = note.start
            end = note.end
            pitch = note.pitch
            velocity = note.velocity
            outputs.append([start, end, pitch, velocity])
    outputs.sort(key=lambda elem: elem[0])
    return outputs



def convert_midis_to_labels(midi_dir, txt_dir): 
    from note_seq import midi_io
    from note_seq import sequences_lib
    
    for midiname in os.listdir(midi_dir):
        suffix = '.'+midiname.split('.')[-1]
        if suffix not in ['.midi', '.mid', '.MID', '.MIDI']:
            continue
        savename = midiname.replace(suffix, '.txt')
        savepath = os.path.join(
            txt_dir, savename).replace('.flac.pred', '')

        midipath = os.path.join(midi_dir, midiname)

        ns = midi_io.midi_file_to_note_sequence(midipath)
        sequence = sequences_lib.apply_sustain_control_changes(ns)
        notes = []
        for note in sequence.notes:
            onset, offset, pitch, velocity = note.start_time, note.end_time, note.pitch, note.velocity
            notes.append([onset, offset, pitch, velocity])
            notes.sort(key=lambda note: note[0])
            
        with open(savepath, 'wt') as f:
            for data in notes:
                onset, offset, pitch, velocity = data
                f.write("{}\t{}\t{}\t{}\n".format(
                        round(onset, 4), round(offset, 4), round(pitch, 4), round(velocity, 4)))

def convert_midis_to_txt(midi_dir, txt_dir=None):
    if txt_dir == None:
        txt_dir = os.path.join(midi_dir, 'txt')
        if not os.path.exists(txt_dir):
            os.makedirs(txt_dir)
    for midiname in os.listdir(midi_dir):
        suffix = '.'+midiname.split('.')[-1]
        if suffix not in ['.midi', '.mid', '.MID', '.MIDI']:
            continue
        savename = midiname.replace(suffix, '.txt')
        savepath = os.path.join(
            txt_dir, savename).replace('.flac.pred', '')

        midipath = os.path.join(midi_dir, midiname)
        datas = extract_labels_from_midi(midipath)

        with open(savepath, 'wt') as f:
            for data in datas:
                onset, offset, pitch, velocity = data
                f.write("{}\t{}\t{}\t{}\n".format(
                        round(onset, 4), round(offset, 4), round(pitch, 4), round(velocity, 4)))

def convert_single_midi_to_txt(midi_path, txt_dir=None):
    if txt_dir == None:
        txt_dir = os.path.join(os.path.dirname(midi_path), 'txt')
        if not os.path.exists(txt_dir):
            os.makedirs(txt_dir)
    midiname = os.path.basename(midi_path)
    suffix = '.'+midiname.split('.')[-1]
    assert suffix in ['.midi', '.mid', '.MID', '.MIDI']
    savename = midiname.replace(suffix, '.txt')
    savepath = os.path.join(txt_dir, savename).replace('.flac.pred', '')

    datas = extract_labels_from_midi(midi_path)

    with open(savepath, 'wt') as f:
        for data in datas:
            onset, offset, pitch, velocity = data
            f.write("{}\t{}\t{}\t{}\n".format(
                    round(onset, 4), round(offset, 4), round(pitch, 4), round(velocity, 4)))


def get_notes(filepath, load_velocity=False):
    notes = []
    filename, suffix = os.path.splitext(os.path.basename(filepath))
    if suffix in ['.midi', '.mid', '.MID', '.MIDI']:
        datas = extract_labels_from_midi(filepath)
        for data in datas:
            onset, offset, pitch, velocity = data
            if load_velocity:
                notes.append([round(float(onset), 4), round(float(offset), 4), round(
                    float(pitch), 4), round(float(velocity), 4)])
            else:
                notes.append([round(float(onset), 4), round(
                    float(offset), 4), round(float(pitch), 4)])
    elif suffix == '.txt':
        with open(filepath) as f:
            for line in f:
                if "OnsetTime" in line:
                    continue
                if load_velocity:
                    onset, offset, pitch, velocity = line.strip().split()
                    notes.append([float(onset), float(offset),
                                 float(pitch), float(velocity)])
                else:
                    onset, offset, pitch = line.strip().split()[:3]
                    notes.append([float(onset), float(offset), float(pitch)])
    else:
        assert False, f'Wrong filepath: {filepath}'

    notes.sort(key=lambda x: x[0])
    return np.array(notes)


def match(labelpath, resultpath, xl_path, onset_tolerance=0.05, log_path=None):
    label = get_notes(labelpath, load_velocity=True)
    result = get_notes(resultpath, load_velocity=True)

    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet(title='sheet1', index=0)

    sheet1.cell(row=1, column=1, value='label')
    sheet1.cell(row=1, column=4, value='result')

    shift = 2
    for i in range(len(label)):
        sheet1.cell(row=i+shift, column=1, value='{:.3f}'.format(label[i, 0]))  # onset
        sheet1.cell(row=i+shift, column=2, value='{:.3f}'.format(label[i, 1]))  # offset
        sheet1.cell(row=i+shift, column=3, value='{:.1f}'.format(label[i, 2])) # pitch
        sheet1.cell(row=i+shift, column=4, value='{:.1f}'.format(label[i, 3])) # velocity

    for i in range(len(result)):
        sheet1.cell(row=i+shift, column=5, value='{:.3f}'.format(result[i, 0]))  # onset(result)
        sheet1.cell(row=i+shift, column=6, value='{:.3f}'.format(result[i, 1]))  # offset(result)
        sheet1.cell(row=i+shift, column=7, value='{:.1f}'.format(result[i, 2])) # pitch(result)
        sheet1.cell(row=i+shift, column=8, value='{:.1f}'.format(result[i, 3])) # velocity(result)

    p, r, f, _ = mir_eval.transcription.precision_recall_f1_overlap(
        label[:, :2], librosa.midi_to_hz(label[:, 2]), 
        result[:, :2], librosa.midi_to_hz(result[:, 2]),
        onset_tolerance=onset_tolerance, offset_ratio=None
    )

    matching_onset = mir_eval.transcription.match_notes(
        label[:, :2], librosa.midi_to_hz(label[:, 2]),
        result[:, :2], librosa.midi_to_hz(result[:, 2]),
        onset_tolerance=onset_tolerance, offset_ratio=None
    )

    orange_fill = PatternFill(fill_type='solid', fgColor="F09B59")
    green_fill = PatternFill(fill_type='solid', fgColor="377D22")
    blue_fill = PatternFill(fill_type='solid', fgColor="3282F6")
    red_fill = PatternFill(fill_type='solid', fgColor="EB3324")
    purple_fill = PatternFill(fill_type='solid', fgColor="732BF5")
    yellow_fill = PatternFill(fill_type='solid', fgColor="FFFF00")
    white_fill = PatternFill(fill_type='solid', fgColor="FFFFFF")

    colors = [orange_fill, green_fill, blue_fill, red_fill, purple_fill]

    for idx, (i, j) in enumerate(matching_onset):
        color = colors[idx % len(colors)]
        sheet1.cell(row=i+shift, column=1).fill = color
        sheet1.cell(row=i+shift, column=2).fill = color
        sheet1.cell(row=i+shift, column=3).fill = color
        sheet1.cell(row=i+shift, column=4).fill = color
        sheet1.cell(row=j+shift, column=5).fill = color
        sheet1.cell(row=j+shift, column=6).fill = color
        sheet1.cell(row=j+shift, column=7).fill = color
        sheet1.cell(row=j+shift, column=8).fill = color

    sheet2 = workbook.create_sheet(title='sheet2', index=0)
    sheet2.cell(row=1, column=1, value='标签')
    sheet2.cell(row=1, column=5, value='结果')
    
    sheet2.cell(row=1, column=20, value='正检')
    sheet2.cell(row=2, column=20, value='{}'.format(len(matching_onset)))
    sheet2.cell(row=1, column=22, value='p')
    sheet2.cell(row=2, column=22, value='{:.2f}'.format(p*100))
    sheet2.cell(row=1, column=23, value='r')
    sheet2.cell(row=2, column=23, value='{:.2f}'.format(r*100))
    sheet2.cell(row=1, column=24, value='f1')
    sheet2.cell(row=2, column=24, value='{:.2f}'.format(f*100))
    
    sheet2.cell(row=3, column=20, value='漏检') # False Positive | Over-Detection
    sheet2.cell(row=4, column=20,value='{}'.format(len(label) - len(matching_onset)))
    
    sheet2.cell(row=5, column=20, value='多检') # False Negative | Under-Detection
    sheet2.cell(row=6, column=20,value='{}'.format(len(result) - len(matching_onset)))
    
    harmonic_FP = 0 # 倍频漏检
    sustain_FP = 0 # 同音漏检
    other_FP = 0 # 其他漏检
    
    matching_onset = np.array(matching_onset)
    matching_label_index = matching_onset[:, 0].tolist()
    matching_result_index = matching_onset[:, 1].tolist()
    # matching[i] == (i, j) where reference note i matches estimated note j.

    for i in range(len(label)):
        if i in matching_label_index:
            sheet2.cell(row=i+shift, column=1, value='{:.3f}'.format(label[i, 0])).fill = orange_fill
            sheet2.cell(row=i+shift, column=2, value='{:.3f}'.format(label[i, 1])).fill = orange_fill
            sheet2.cell(row=i+shift, column=3, value='{:.1f}'.format(label[i, 2])).fill = orange_fill
            sheet2.cell(row=i+shift, column=4, value='{:.1f}'.format(label[i, 3])).fill = orange_fill
        else: # 漏检
            sheet2.cell(row=i+shift, column=1, value='{:.3f}'.format(label[i, 0]))
            sheet2.cell(row=i+shift, column=2, value='{:.3f}'.format(label[i, 1]))
            sheet2.cell(row=i+shift, column=3, value='{:.1f}'.format(label[i, 2]))
            sheet2.cell(row=i+shift, column=4, value='{:.1f}'.format(label[i, 3]))
            got_reason = False

            for j in range(max(0, i-9), min(len(label)-1, i+9)):  # 最多同时10个音
                # onset 在其他倍频(二倍或三倍)音符时间内
                if label[i, 0] >= label[j, 0] - onset_tolerance and label[i, 0] <= label[j, 1] + onset_tolerance:
                    if abs(abs(label[j, 2] - label[i, 2]) % 12) <= HARMONIC_TOLERANCE or \
                            abs(abs(label[j, 2] - label[i, 2]) % 19) <= HARMONIC_TOLERANCE:
                        harmonic_FP += 1
                        sheet2.cell(row=i+shift, column=3).fill = red_fill
                        sheet2.cell(row=j+shift, column=3).fill = green_fill
                        got_reason = True
                        break
            if not got_reason:
                for j in range(i - 1, 0, -1): # 同一个音短时间内第二次按下导致漏检
                    if label[i, 0] - label[j, 1] > onset_tolerance * SUSTAIN_TOLERANCE:
                        break
                    elif label[j, 2] == label[i, 2]:
                        sustain_FP += 1
                        sheet2.cell(row=j+shift, column=3).fill = yellow_fill
                        sheet2.cell(row=i+shift, column=3).fill = yellow_fill
                        sheet2.cell(row=i+shift, column=7, value='{:.3f}'.format(label[i, 0] - label[j, 1])).fill = yellow_fill
                        break
            if not got_reason:
                other_FP += 1

    harmonic_FN = 0 # 倍频多检
    sustain_FN = 0 # 同音多检
    other_FN = 0 # 其他多检
    shift_FN = 0
    label_index = 0
    for i in range(len(result)):
        if i in matching_result_index:
            matching_index = matching_result_index.index(i)
            label_index = matching_label_index[matching_index]
            shift_FN = 0
            sheet2.cell(row=label_index+shift, column=5, value='{:.3f}'.format(result[i, 0])).fill = green_fill
            sheet2.cell(row=label_index+shift, column=6, value='{:.3f}'.format(result[i, 1])).fill = green_fill
            sheet2.cell(row=label_index+shift, column=7, value='{:.1f}'.format(result[i, 2])).fill = green_fill
            sheet2.cell(row=label_index+shift, column=8, value='{:.1f}'.format(result[i, 3])).fill = green_fill
        else:
            shift_FN += 1
            sheet2.cell(row=label_index+shift, column=5 + 4 * shift_FN, value='{:.3f}'.format(result[i, 0]))
            sheet2.cell(row=label_index+shift, column=6 + 4 * shift_FN, value='{:.3f}'.format(result[i, 1]))
            sheet2.cell(row=label_index+shift, column=7 + 4 * shift_FN, value='{:.1f}'.format(result[i, 2]))
            sheet2.cell(row=label_index+shift, column=8 + 4 * shift_FN, value='{:.1f}'.format(result[i, 3]))
            got_reason = False
            for j in range(max(0, label_index-9), min(len(label)-1, label_index+9)):  # 10 fingers
                if result[i, 0] >= label[j, 0] - onset_tolerance and result[i, 0] <= label[j, 1] + onset_tolerance:
                    if abs(abs(result[i, 2] - label[j, 2]) % 12) <= HARMONIC_TOLERANCE or \
                            abs(abs(result[i, 2] - label[j, 2]) % 19) <= HARMONIC_TOLERANCE:
                        harmonic_FN += 1
                        got_reason = True
                        sheet2.cell(row=j+shift, column=3).fill = white_fill
                        break
            if not got_reason:
                for j in range(min(len(label)-1, label_index+1), 0, -1):
                    if result[i, 0] - label[j, 1] > onset_tolerance * SUSTAIN_TOLERANCE:
                        break
                    if result[i, 2] == label[j, 2]:
                        sustain_FN += 1
                        sheet2.cell(row=j+shift, column=3).fill = white_fill
                        break
            if not got_reason:
                other_FN += 1


    sheet2.cell(row=3, column=22, value='倍频漏检').fill = orange_fill
    sheet2.cell(row=4, column=22, value=harmonic_FP).fill = orange_fill
    sheet2.cell(row=3, column=23, value='同音漏检').fill = orange_fill
    sheet2.cell(row=4, column=23, value=sustain_FP).fill = orange_fill
    sheet2.cell(row=3, column=24, value='其他漏检').fill = orange_fill
    sheet2.cell(row=4, column=24, value=other_FP).fill = orange_fill
    
    sheet2.cell(row=5, column=22, value='倍频多检')
    sheet2.cell(row=6, column=22, value=harmonic_FN)
    sheet2.cell(row=5, column=23, value='同音多检')
    sheet2.cell(row=6, column=23, value=sustain_FN)
    sheet2.cell(row=5, column=24, value='其他多检')
    sheet2.cell(row=6, column=24, value=other_FN)

    if log_path:
        print(xl_path)
        prf_str = '[Onset]\tPrecision: {:.2f}% Recall: {:.2f}% F1: {:.2f}% \t正检: {}'.format(p * 100, r * 100, f * 100, len(matching_onset))
        error_str = '漏检: {}\t=>  倍频漏检: {}\t同音漏检: {}\t其他漏检: {}\n'.format(len(label)  - len(matching_onset), harmonic_FP, sustain_FP, other_FP) + \
                    '多检: {}\t=>  倍频多检: {}\t同音多检: {}\t其他多检: {}\n'.format(len(result) - len(matching_onset), harmonic_FN, sustain_FN, other_FN)

        print(prf_str + '\n' + error_str)
        with open(log_path, 'a') as log_file:
            log_file.write(prf_str + '\n' + error_str)

    p, r, f, _ = mir_eval.transcription.precision_recall_f1_overlap(
        label[:, :2], librosa.midi_to_hz(label[:, 2]), 
        result[:, :2], librosa.midi_to_hz(result[:, 2]),
        onset_tolerance=onset_tolerance
    )

    matching_offset = mir_eval.transcription.match_notes(
        label[:, :2], librosa.midi_to_hz(label[:, 2]),
        result[:, :2], librosa.midi_to_hz(result[:, 2]),
        onset_tolerance=onset_tolerance
    )
    matching_offset = np.array(matching_offset)
    matching_offset_index = matching_offset[:, 1].tolist()
    long_offset = 0
    harmonic_long = 0 # 倍频过长
    sustain_long = 0 # 同音过长
    other_long = 0 # 其他过长
    
    short_offset = 0
    harmonic_short = 0 # 倍频过短
    sustain_short = 0 # 同音过短
    other_short = 0 # 其他过短
    
    for i in matching_result_index:
        if i not in matching_offset_index:
            matching_index = matching_result_index.index(i)
            label_index = matching_label_index[matching_index]
            
            if result[i, 1] > label[label_index, 1]:
                # assert result[i,2] == label[i,2] and result[i, 1] - label[label_index, 1] > onset_tolerance, f"result:{result[i,:]}\tlabel:{label[label_index,:]}"
                long_offset += 1
                sheet2.cell(row=label_index+shift, column=2, value='{:.3f}'.format(label[label_index, 1])).fill = yellow_fill
                got_reason = False
                for j in range(max(0, label_index-9), min(len(label)-1, label_index+9)):  # 10 fingers
                    if result[i, 1] >= label[j, 0] - onset_tolerance and result[i, 1] <= label[j, 1] + onset_tolerance:
                        if abs(abs(result[i, 2] - label[j, 2]) % 12) <= HARMONIC_TOLERANCE or \
                                abs(abs(result[i, 2] - label[j, 2]) % 19) <= HARMONIC_TOLERANCE:
                            harmonic_long += 1
                            got_reason = True
                            break
                if not got_reason:
                    for j in range(min(len(label)-1, label_index+1), len(label)):
                        if label[j, 0] - result[i, 1] > onset_tolerance * SUSTAIN_TOLERANCE:
                            break
                        if result[i, 2] == label[j, 2]:
                            sustain_long += 1
                            break
                if not got_reason:
                    other_long += 1
            else:
                # assert result[i,2] == label[i,2] and label[label_index, 1] - result[i, 1] > onset_tolerance, f"result:{result[i,:]}\tlabel:{label[label_index,:]}"
                short_offset += 1
                sheet2.cell(row=label_index+shift, column=2, value='{:.3f}'.format(label[label_index, 1])).fill = purple_fill
                got_reason = False
                for j in range(max(0, label_index-9), min(len(label)-1, label_index+9)):  # 10 fingers
                    if result[i, 1] >= label[j, 0] - onset_tolerance and result[i, 1] <= label[j, 1] + onset_tolerance:
                        if abs(abs(result[i, 2] - label[j, 2]) % 12) <= HARMONIC_TOLERANCE or \
                                abs(abs(result[i, 2] - label[j, 2]) % 19) <= HARMONIC_TOLERANCE:
                            harmonic_short += 1
                            got_reason = True
                            break
                if not got_reason:
                    for j in range(min(len(label)-1, label_index+1), len(label)):
                        if label[j, 0] - result[i, 1] > onset_tolerance * SUSTAIN_TOLERANCE:
                            break
                        if result[i, 2] == label[j, 2]:
                            sustain_short += 1
                            break
                if not got_reason:
                    other_short += 1

    if log_path:
        prf_str = '[Offset]\tPrecision: {:.2f}% Recall: {:.2f}% F1: {:.2f}%\t正检: {}'.format(p * 100, r * 100, f * 100, len(matching_offset))
        error_str = '过长: {}\t=> 倍频过长: {}\t同音过长: {}\t其他过长: {}\n'.format(long_offset, harmonic_long, sustain_long, other_long) + \
                    '过短: {}\t=> 倍频过短: {}\t同音过短: {}\t其他过短: {}'.format(short_offset, harmonic_short, sustain_short, other_short)

        print(prf_str + '\n' + error_str + '\n\n')
        with open(log_path, 'a') as log_file:
            log_file.write(prf_str + '\n' + error_str + '\n\n')

    workbook.save(filename=xl_path)
    workbook.close()
    return [len(matching_onset), len(label) - len(matching_onset), len(result) - len(matching_onset), 
            harmonic_FP, sustain_FP, other_FP, harmonic_FN, sustain_FN, other_FN, 
            len(matching_offset), long_offset, short_offset, 
            harmonic_long, sustain_long, other_long,
            harmonic_short, sustain_short, other_short]


if __name__ == '__main__':
    # convert_midis_to_txt('/home/data/wrm_data/temp_task/piano_transcription-master/workspaces/probs/Four_blocks/none/dataset=maestro/test/batch_size=2/532500_iterations/532500_iterations_midi/','/home/data/wrm_data/temp_task/piano_transcription-master/workspaces/probs/Four_blocks/none/dataset=maestro/test/batch_size=2/532500_iterations/532500_iterations_midi/txt')
    # convert_midis_to_labels("/home/data/MAPS/MapsMus/ENSTDkAm/MUS/", "/home/data/wrm_data/temp_task/piano_transcription-master/workspaces/hdf5s/maps_labels")
    # convert_midis_to_labels("/home/data/wrm_data/temp_task/piano_transcription-master/piano_transcription_hFT-Transformer-master/corpus/MAESTRO-V3/maestro-v3.0.0/2006", "/home/data/wrm_data/temp_task/piano_transcription-master/workspaces/hdf5s/Maestro_0")
    args = parse_args()
    label_dir = args.label_dir
    result_dir = args.result_dir
    # name = args.name
    save_dir = args.save_dir
    if save_dir is None:
        save_dir = os.path.join(result_dir, 'xlsx')
        os.makedirs(save_dir, exist_ok=True)

    log_path = os.path.join(save_dir, datetime.now().strftime('%y%m%d') + '_match.log')
    if os.path.exists(log_path):
        os.remove(log_path)

    filenames = os.listdir(result_dir)
    errors = np.zeros(18)
    for i, name in enumerate(filenames):
        prefix, suffix = os.path.splitext(name)
        if suffix not in ['.txt', '.midi', '.mid', '.MID', '.MIDI']:
            print(f"Unsupported file format and pass: {name}")
            continue
        else:
            print(f"{i}\t|\tCounting errors in {name}", end='\r')
            save_path = os.path.join(save_dir, name.replace(suffix, '.xlsx'))
            with open(log_path, 'a') as log_file:
                log_file.write(f"{i}\t|\t{log_path}\n")
        result_path = os.path.join(result_dir, name)
        for suffix in ['.txt', '.midi', '.mid', '.MID', '.MIDI']:
            label_path = os.path.join(label_dir, prefix + suffix)
            if os.path.exists(label_path):
                break
        errors += match(label_path, result_path, save_path, args.onset_win, log_path)


    error_str = 'SUSTAIN_TOLERANCE_TIME = {}s\tHARMONIC_TOLERANCE = {}\n'.format(SUSTAIN_TOLERANCE * args.onset_win, HARMONIC_TOLERANCE) + \
                'Onset  正检: {} \n漏检: {}\t=>  倍频漏检: {}\t同音漏检: {}\t其他漏检: {}\n'.format(errors[0],errors[1],errors[3],errors[4],errors[5]) + \
                '多检: {}\t=>  倍频多检: {}\t同音多检: {}\t其他多检: {}\n'.format(errors[2],errors[6],errors[7],errors[8]) + \
                'Offset 正检: {}\n过长: {}\t=> 倍频过长: {}\t同音过长: {}\t其他过长: {}\n'.format(errors[9],errors[10],errors[12],errors[13],errors[14]) + \
                '过短: {}\t=> 倍频过短: {}\t同音过短: {}\t其他过短: {}'.format(errors[11],errors[15],errors[16],errors[17])

    print(error_str)
    print(f'Saved at {save_dir}')
    with open(log_path, 'a') as log_file:
        log_file.write(error_str)
